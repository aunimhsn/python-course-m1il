from openpyxl import load_workbook

# 1. Charger / Importer le fichier Excel
wb = load_workbook('transactions.xlsx')
ws = wb.active

# 2. Lire les cellules de c2 à cN (c4)
# 3. Parser chacune des cellules c2 à cN pour obtenir un float
for i in range(2, ws.max_row + 1):
    price = ws[f'C{i}'].value
    price_float = float(price.replace('€', '').replace(',', '.'))

    discount_price = round(price_float * 0.9, 2)

    # 4. Ecrire dans la colonne D la réduction
    ws[f'D{i}'].value = discount_price

wb.save('transactions_discount.xlsx')
